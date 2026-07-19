"""
build_report.py  -  Compile one Excel workbook for the day + running total.

Reads the raw poll data directly (robust; independent of the rollup step),
classifies each station using the current master list tags (funded / multi-
family), and writes:

  data/reports/EV_Bellevue_Report_<date>.xlsx   two tabs: "Today", "Running Total"
  data/reports/_email_body.md                   the text used in the nightly email

Method notes:
  Occupancy % = occupied / (available + occupied + reserved). Broken
  (out-of-service) and unknown ports are excluded from that fraction.
  A charger that was OUT OF SERVICE for the whole day (no usable reading at all)
  is reported separately as "offline", NOT counted as "zero-use" — a broken
  charger is not the same as an idle one.
  "Zero-use" = a charger that had usable readings but was never seen occupied.

Target day: if no date is given, report the most recent COMPLETED day. This makes
a late-running nightly job (e.g. delayed to 2am) still report yesterday's full
data instead of an empty "today".
"""

import csv
import json
import sys
from collections import defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
CONFIG = yaml.safe_load((ROOT / "config.yaml").read_text())
TZ = ZoneInfo(CONFIG["project"]["timezone"])
RAW_DIR = ROOT / "data" / "raw"
REPORTS = ROOT / "data" / "reports"
REPORTS.mkdir(parents=True, exist_ok=True)

PILOT_START = date.fromisoformat(CONFIG["pilot"]["start_date"])
PILOT_END = date.fromisoformat(CONFIG["pilot"]["end_date"])
EXPECTED_CYCLES = 28   # ~ every 30 min, 7am-9pm
HEALTHY_MIN = 20       # fewer than this = flag the day as "thin"

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def load_master_tags():
    p = ROOT / "data" / "master" / "stations.json"
    tags = {}
    if p.exists():
        for s in json.loads(p.read_text()):
            tags[s["nrel_id"]] = {
                "address": s.get("address", ""),
                "network": s.get("network", ""),
                "funded": bool(s.get("funded")),
                "multifamily": bool(s.get("multifamily")),
            }
    return tags


def occ(row):
    a, o, r = int(row["available"]), int(row["occupied"]), int(row["reserved"])
    d = a + o + r
    return (o / d) if d > 0 else None


def read_day(day_str):
    p = RAW_DIR / f"{day_str}.csv"
    if not p.exists():
        return []
    with p.open() as f:
        return list(csv.DictReader(f))


def poll_cycles(rows):
    return len({r["timestamp_local"] for r in rows})


def raw_dates():
    out = []
    for p in RAW_DIR.glob("*.csv"):
        try:
            date.fromisoformat(p.stem)
            out.append(p.stem)
        except ValueError:
            continue
    return sorted(out)


def summarize_day(rows, tags):
    per = defaultdict(lambda: {"occ_sum": 0.0, "n": 0, "ever": False,
                               "ports": 0, "network": "", "id": ""})
    for row in rows:
        sid = row["nrel_id"]
        st = per[sid]
        st["id"] = sid
        f = occ(row)
        if f is not None:
            st["occ_sum"] += f
            st["n"] += 1
        if int(row["occupied"]) > 0:
            st["ever"] = True
        st["ports"] = max(st["ports"], int(row["total_ports"]))
        st["network"] = tags.get(sid, {}).get("network") or row["network"]
    out = {}
    for sid, st in per.items():
        usable = st["n"] > 0
        avg = (100 * st["occ_sum"] / st["n"]) if usable else None
        out[sid] = {
            "id": sid,
            "network": st["network"],
            "address": tags.get(sid, {}).get("address", ""),
            "ports": st["ports"],
            "samples": st["n"],
            "avg_occ": avg,
            "offline": not usable,                       # no usable reading all day
            "zero_use": usable and not st["ever"],       # had data, never occupied
            "funded": tags.get(sid, {}).get("funded", False),
            "multifamily": tags.get(sid, {}).get("multifamily", False),
        }
    return out


def block_stats(stations):
    """Aggregate, separating offline (broken) chargers from usable ones."""
    usable = [s for s in stations if not s["offline"]]
    offline = [s for s in stations if s["offline"]]
    vals = [s["avg_occ"] for s in usable if s["avg_occ"] is not None]
    avg = round(sum(vals) / len(vals), 1) if vals else None
    zero = sum(1 for s in usable if s["zero_use"])
    return {
        "total": len(stations),
        "usable": len(usable),
        "offline": len(offline),
        "ports": sum(s["ports"] for s in stations),
        "avg_occ": avg,
        "zero_ct": zero,
        "zero_pct": round(100 * zero / len(usable), 1) if usable else None,
    }


def sheet_header(ws, title, subtitle):
    ws.append([title]); ws["A1"].font = Font(bold=True, size=14)
    ws.append([subtitle]); ws["A2"].font = Font(italic=True, color="555555")
    ws.append([])


def style_table_header(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT


def pick_target(explicit):
    if explicit:
        return date.fromisoformat(explicit)
    now = datetime.now(TZ).date()
    dated = raw_dates()
    if not dated:
        return now
    ns = now.isoformat()
    # Use today only if it already looks like a full day; otherwise the most
    # recent completed prior day. Prevents empty/misdated reports from late runs.
    if ns in dated and poll_cycles(read_day(ns)) >= HEALTHY_MIN:
        return now
    prior = [d for d in dated if d < ns]
    return date.fromisoformat(prior[-1] if prior else dated[-1])


def health_label(cycles):
    return "healthy" if cycles >= HEALTHY_MIN else "THIN — fewer samples than expected"


def main(report_date_str=None):
    tags = load_master_tags()
    today = pick_target(report_date_str)

    days = []
    d = PILOT_START
    while d <= today:
        days.append(d.isoformat())
        d += timedelta(days=1)
    if not days:
        days = [today.isoformat()]

    day_rows = {ds: read_day(ds) for ds in days}
    day_summaries = {ds: summarize_day(day_rows[ds], tags) for ds in days}
    day_cycles = {ds: poll_cycles(day_rows[ds]) for ds in days}

    wb = Workbook()
    tstr = today.isoformat()
    dow = today.strftime("%A")
    cyc_today = day_cycles.get(tstr, 0)

    # ================= TODAY =================
    ws = wb.active
    ws.title = "Today"
    pilot_day = (today - PILOT_START).days + 1
    pilot_note = (f"Pilot day {pilot_day} of 7" if PILOT_START <= today <= PILOT_END
                  else ("Pre-pilot test" if today < PILOT_START else "Post-pilot (still collecting)"))
    sheet_header(ws, f"Bellevue EV Charger Usage — {tstr} ({dow})",
                 f"{pilot_note}. Occupancy = share of ports in use, 7am-9pm, 30-min polls.")

    today_stations = list(day_summaries.get(tstr, {}).values())
    s = block_stats(today_stations)
    funded = [x for x in today_stations if x["funded"]]
    mf = [x for x in today_stations if x["multifamily"]]

    ws.append(["Headline (today)", ""])
    ws.append(["Chargers polled", s["total"]])
    ws.append(["  of which reporting usable data", s["usable"]])
    ws.append(["  of which offline all day (excluded)", s["offline"]])
    ws.append(["Total charging ports", s["ports"]])
    ws.append(["Average occupancy (of usable chargers)",
               f"{s['avg_occ']}%" if s["avg_occ"] is not None else "n/a"])
    ws.append(["Chargers with ZERO use today",
               f"{s['zero_ct']} of {s['usable']} usable ({s['zero_pct']}%)"])
    ws.append(["Data quality — poll cycles today",
               f"{cyc_today} of ~{EXPECTED_CYCLES} expected ({health_label(cyc_today)})"])
    ws.append([])
    ws.append(["Subsets", "Chargers", "Avg occupancy", "Zero-use"])
    for label, grp in [("State-funded", funded), ("Apartment/condo", mf)]:
        if grp:
            b = block_stats(grp)
            ws.append([label, b["usable"], f"{b['avg_occ']}%", f"{b['zero_pct']}%"])
        else:
            ws.append([label, 0, "pending NREL", "pending NREL"])
    ws.append([])

    hdr_row = ws.max_row + 1
    ws.append(["Network", "Address", "Ports", "Samples", "Avg occupancy %",
               "Zero-use?", "Offline?", "Funded", "Apartment/condo"])
    style_table_header(ws, hdr_row, 9)
    for st in sorted(today_stations, key=lambda x: (x["avg_occ"] is None, x["avg_occ"] or 0)):
        ws.append([st["network"], st["address"], st["ports"], st["samples"],
                   round(st["avg_occ"], 1) if st["avg_occ"] is not None else "n/a",
                   "YES" if st["zero_use"] else "", "YES" if st["offline"] else "",
                   "YES" if st["funded"] else "", "YES" if st["multifamily"] else ""])
    for col, w in zip("ABCDEFGHI", [16, 40, 7, 9, 15, 10, 9, 9, 15]):
        ws.column_dimensions[col].width = w

    # ================= RUNNING TOTAL =================
    ws2 = wb.create_sheet("Running Total")
    sheet_header(ws2, f"Running Total — {CONFIG['pilot']['label']}",
                 f"Cumulative {PILOT_START.isoformat()} to {tstr}. "
                 f"Official window {PILOT_START.isoformat()}-{PILOT_END.isoformat()}.")

    all_station_days = [st for ds in days for st in day_summaries[ds].values()]
    cum = block_stats(all_station_days)
    uniq = len({st["id"] for st in all_station_days})
    ws2.append(["Days collected so far", sum(1 for ds in days if day_summaries[ds])])
    ws2.append(["Unique chargers seen", uniq])
    ws2.append(["Cumulative average occupancy",
                f"{cum['avg_occ']}%" if cum["avg_occ"] is not None else "n/a"])
    ws2.append(["Usable station-days with zero use",
                f"{cum['zero_ct']} of {cum['usable']} ({cum['zero_pct']}%)"])
    ws2.append([])

    hdr2 = ws2.max_row + 1
    ws2.append(["Date", "Day", "Chargers", "Avg occupancy %",
                "% zero-use", "Poll cycles", "Data quality"])
    style_table_header(ws2, hdr2, 7)
    for ds in days:
        sd = list(day_summaries[ds].values())
        if not sd:
            continue
        b = block_stats(sd)
        ws2.append([ds, date.fromisoformat(ds).strftime("%a"), b["usable"],
                    b["avg_occ"], b["zero_pct"], day_cycles[ds],
                    "healthy" if day_cycles[ds] >= HEALTHY_MIN else "thin"])
    for col, w in zip("ABCDEFG", [12, 6, 10, 16, 12, 12, 14]):
        ws2.column_dimensions[col].width = w

    out_path = REPORTS / f"EV_Bellevue_Report_{tstr}.xlsx"
    wb.save(out_path)

    # ---- email body ----
    lines = [
        f"# Bellevue EV Charger Pilot — {tstr} ({dow})",
        f"*{pilot_note}*", "",
        "## Today",
        f"- Chargers polled: **{s['total']}** ({s['usable']} usable, {s['offline']} offline all day)",
        (f"- Average occupancy (usable): **{s['avg_occ']}%**" if s["avg_occ"] is not None
         else "- Average occupancy: n/a"),
        f"- Chargers with zero use: **{s['zero_ct']} of {s['usable']} ({s['zero_pct']}%)**",
        f"- Data quality: **{cyc_today} of ~{EXPECTED_CYCLES}** poll cycles ({health_label(cyc_today)})",
        "", "## Running total (pilot to date)",
        f"- Days collected: **{sum(1 for ds in days if day_summaries[ds])}**",
        (f"- Cumulative average occupancy: **{cum['avg_occ']}%**" if cum["avg_occ"] is not None
         else "- Cumulative average occupancy: n/a"),
        f"- Usable station-days with zero use: **{cum['zero_pct']}%**", "",
    ]
    if not funded and not mf:
        lines.append("_Funded and apartment/condo subsets activate automatically "
                     "once the NREL database is reachable again._")
    lines.append("")
    lines.append("Full detail is in the attached spreadsheet (Today + Running Total tabs).")
    (REPORTS / "_email_body.md").write_text("\n".join(lines))

    print(f"Report written: {out_path.name} (target day {tstr}, {cyc_today} cycles)")
    print(f"Today: {s['usable']} usable chargers, avg occupancy {s['avg_occ']}%, "
          f"{s['zero_pct']}% zero-use, {s['offline']} offline")
    return out_path


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else None)
